# -*- coding: utf-8 -*-
"""
make_weekly_performance.py ── 주간성과 요약 (보유자산 게시판 하위탭 "주간성과")
──────────────────────────────────────────────────────────────────────────
목적: 여러 자동매매 봇(주문 스타일)의 성과를 주(월~금) 단위로 비교.
      "뭐가 뛰어나고 뭐가 안좋은지" 를 한눈에.

데이터 소스: D:\\py\\0order\\intraday_signals\\*.json  (모든 봇 체결이 통합·git동기화된 유일한 소스)
  - 기존 장부(정리본 CSV, danta JSON)는 건드리지 않는다. 이 스트림만 읽어 재구성.
  - make_danta_journal.py 와 동일한 fill 이벤트 스키마 사용.

봇 귀속(핵심):
  - 계좌(silo)별로 포지션 lifecycle(flat→매수…→flat) 를 재구성.
  - 매수가 포지션의 '소유 봇' 을 확정하고, 매도는 그 포지션에 흡수(FIFO/평균단가).
    → rocket 이 사고 출근 자동매도가 팔아도(매도 태그=leader) 매수 소유봇(rocket)에 정확히 귀속.
  - 매도 자체 태그에 의존하지 않으므로 다계좌·수동 분할매도도 소유봇 성과로 합산됨.

봇 매핑:
  (1887, rocket)          → ROCKET      (스윙 / 보유=일 / 000_6rocket.bat)
  (1887, leader/그외태그) → 주도주       (스윙 / 보유=일)
  (1887DIP, *)            → 저점사다리   (당일 / 보유=분 / 0_5min_chu_low_sam_hynix_lowbuy.py, ~260731 이전)
  (8042DIP, *)            → 저점사다리   (당일 / 보유=분 / 0_5min_chu_low_sam_hynix_lowbuy.py, 260731~ 계좌이관)
  (8042CHU, *)            → 2X단타       (당일 / 보유=분)
  (2773, *)               → 5min저고단타 (당일 / 보유=분)
  (8042, *)               → 기타(8042)   (분류불가 수동성)
  ※ (8042, lowhigh) 는 실험용(며칠 전 2773 로 이관) → 통계에서 완전 제외.

교차기록 중복제거(1887DIP / 8042DIP):
  0_5min_chu_low_sam_hynix_lowbuy.py 는 자기 체결을 acct="1887DIP"(260731 계좌이관 후 "8042DIP")
  로 직접 기록하는데, sync_1887_fills.py/sync_8042_fills.py 가 나중에 ka10076 로 해당 계좌
  전체를 재조회하면서 같은 체결을 (태깅 안 됨 → 기본값) acct="1887"/"8042" 로 또 기록해
  이중계상이 생긴다.
  → load_fills() 에서 같은 날 acct="1887DIP"(또는 "8042DIP") 에 있는 ord_no 와 겹치는
    acct="1887"(또는 "8042") fill 은 버리고 "*DIP" 쪽(정확한 전략태그)만 남긴다.

산출:
  - report-us/weekly_performance.json   (holdings.html 주간성과 탭이 fetch)
  - report-us/weekly_performance.csv    (Excel 호환, utf-8-sig)
  - report-us/weekly_performance.xlsx   (openpyxl 있을 때만)

실행: python -X utf8 make_weekly_performance.py   (cwd: D:\\py\\report-us)
"""
import os
import re
import csv
import sys
import glob
import json
from datetime import datetime, date, timedelta
from collections import defaultdict

BASE_DIR   = r"D:\py"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)
import trade_costs  # 수수료·증권거래세 요율 단일 원천
SIGNAL_DIR = os.path.join(BASE_DIR, "0order", "intraday_signals")
TR_LEDGER_DIR = os.path.join(BASE_DIR, "0order", "tr", "ledger")
OUT_DIR      = SCRIPT_DIR
OUT_JSON      = os.path.join(OUT_DIR, "weekly_performance.json")       # 웹페이지 피드(최근 N주)
OUT_JSON_FULL = os.path.join(OUT_DIR, "weekly_performance_full.json")  # 전체 누적(분석용)
OUT_CSV       = os.path.join(OUT_DIR, "weekly_performance.csv")        # 전체 누적
OUT_XLSX      = os.path.join(OUT_DIR, "weekly_performance.xlsx")       # 전체 누적

# 웹페이지 표시 구간: 최근 8주(약 2달). 데이터 파일(full json/csv/xlsx)은 전체 보관.
RECENT_WEEKS = 8

# 미국 체결가는 달러, 한국은 원. 금액 합계를 내려면 통화를 맞춰야 한다.
#
# ★ 주차별 환율 스냅샷(복기 재현성)
#   예전엔 "실행 시점의 최신 USDKRW 하나"로 전 기간을 일괄 환산했다. 그러면 몇 달 뒤
#   다시 돌릴 때 과거 주차의 원화 금액이 통째로 바뀌어서 복기가 불가능하다.
#   → 주(월요일) 단위로 실제 쓴 환율을 weekly_fx_snapshot.json 에 기록하고,
#     한 번 기록된 과거 주차는 재실행해도 그 값을 그대로 재사용(=동결)한다.
#     진행 중인 이번 주만 매 실행 갱신되고, 주가 끝나면 자연히 동결된다.
#   ※ %지표(승률/PF/기대값%)는 통화무관이라 환율과 무관하게 원래부터 안정적.
EQUITY_CSV = os.path.join(SCRIPT_DIR, "equity_daily.csv")
FX_SNAPSHOT_JSON = os.path.join(SCRIPT_DIR, "weekly_fx_snapshot.json")
USDKRW_FALLBACK = 1400.0


def _equity_usdkrw_by_date():
    """equity_daily.csv → {날짜(YYYY-MM-DD): usdkrw}. 계좌가 여러 행이면 마지막 값."""
    out = {}
    try:
        with open(EQUITY_CSV, encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                d = (r.get("date") or "").strip()
                try:
                    v = float(r.get("usdkrw") or 0)
                except ValueError:
                    continue
                if d and v > 0:
                    out[d] = v
    except Exception:
        pass
    return out


def _latest_usdkrw():
    try:
        with open(EQUITY_CSV, encoding="utf-8-sig") as f:
            rates = [float(r["usdkrw"]) for r in csv.DictReader(f)
                     if str(r.get("account", "")).isdigit() and float(r.get("usdkrw") or 0) > 0]
        if rates:
            return rates[-1]
    except Exception:
        pass
    # equity_daily.csv 가 아직 없을 때만 fx_rate.json(공용 캐시) → 그것도 없으면 상수.
    try:
        from fx_rate import get_usdkrw
        return get_usdkrw(USDKRW_FALLBACK)
    except Exception:
        return USDKRW_FALLBACK


USDKRW = _latest_usdkrw()          # "오늘자" 환율 (신규/진행중 주차에만 쓰임)
_EQUITY_FX = _equity_usdkrw_by_date()
_FX_SNAP = {"rates": {}, "meta": {}}   # load_fx_snapshot() 이 채운다
_FX_DIRTY = False


def load_fx_snapshot():
    """weekly_fx_snapshot.json 로드. 없으면 빈 상태로 시작(첫 실행 시 자동 생성)."""
    global _FX_SNAP
    try:
        with open(FX_SNAPSHOT_JSON, encoding="utf-8") as f:
            d = json.load(f)
        _FX_SNAP = {"rates": d.get("rates") or {}, "meta": d.get("meta") or {}}
    except Exception:
        _FX_SNAP = {"rates": {}, "meta": {}}
    return _FX_SNAP


def _week_rate_from_equity(monday):
    """그 주(월~금) equity_daily.csv 환율 중 가장 늦은 날 값. 없으면 None."""
    for i in (4, 3, 2, 1, 0):
        v = _EQUITY_FX.get((monday + timedelta(days=i)).isoformat())
        if v:
            return v, "equity_daily"
    return None, None


def fx_for_week(monday):
    """해당 주(월요일 date)의 USD→KRW 환산율. 과거 주차는 스냅샷 고정값."""
    global _FX_DIRTY
    wk = monday.isoformat()
    is_current = (monday == week_monday(date.today()))
    if not is_current and wk in _FX_SNAP["rates"]:
        return float(_FX_SNAP["rates"][wk])          # 동결 — 재실행해도 안 바뀜
    rate, src = _week_rate_from_equity(monday)
    if rate is None:
        rate, src = USDKRW, "latest_fallback"        # 그 주 환율 기록이 없던 과거 구간
    prev = _FX_SNAP["rates"].get(wk)
    if prev is None or float(prev) != float(rate):
        _FX_SNAP["rates"][wk] = round(float(rate), 4)
        _FX_SNAP["meta"][wk] = {"src": src, "written": date.today().isoformat(),
                                "status": "open" if is_current else "frozen"}
        _FX_DIRTY = True
    elif is_current:
        _FX_SNAP["meta"].setdefault(wk, {})["status"] = "open"
    return float(rate)


def save_fx_snapshot():
    if not _FX_DIRTY:
        print(f"  FX 스냅샷 변경 없음 ({len(_FX_SNAP['rates'])}주 고정)")
        return
    # 주가 끝난 주차는 status 를 frozen 으로 확정
    cur = week_monday(date.today()).isoformat()
    for wk, m in _FX_SNAP["meta"].items():
        if wk != cur:
            m["status"] = "frozen"
    payload = {
        "_note": "주간성과 원화 환산에 쓴 주차별 USDKRW. 과거 주차는 재실행해도 고정 "
                 "(복기 시 원화 금액 재현용). status=open 인 이번 주만 갱신된다.",
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "rates": dict(sorted(_FX_SNAP["rates"].items())),
        "meta": dict(sorted(_FX_SNAP["meta"].items())),
    }
    with open(FX_SNAPSHOT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"  OK {os.path.relpath(FX_SNAPSHOT_JSON, OUT_DIR)} ({len(payload['rates'])}주)")

# ───────── 코인(업비트) 주간(월~일) 실현손익 설정 ─────────
# 주식봇과 달리 코인은 주말도 거래되므로 월~일 7일 기준.
# 실현손익 = MTF 본전략(평균단가 replay, net) + scalp(realized_pnl 컬럼) 합산.
COIN_FEE = 0.0005  # 업비트 KRW마켓 수수료 0.05% (reserved_fee 로 확인). net 계산에 매수·매도 양쪽 적용.
COIN_MTF_CSV = os.path.join(BASE_DIR, "coin", "0order", "0_upbit_btc_eth_mtf_trades_v3_live.csv")
COIN_SCALP_CSVS = [  # 우선순위: v3 있으면 v3, 없으면 v2
    os.path.join(BASE_DIR, "coin", "0order", "0_upbit_scalp_trades_v3_live.csv"),
    os.path.join(BASE_DIR, "coin", "0order", "0_upbit_scalp_trades_v2_live.csv"),
]
COIN_MARKETS = {"KRW-BTC": "btc", "KRW-ETH": "eth"}
COIN_WEEKDAYS = ["월", "화", "수", "목", "금", "토", "일"]

# ───────── Binance USDT-M 선물 주간(월~일) 실현손익 설정 ─────────
# 코인과 같은 이유로 별도 섹션: 수량이 소수(0.043 ETH 등)라 intraday_signals/TR ledger의
# int() 수량 전제 파이프라인에 못 넣는다. 실현손익도 로컬 주문로그가 아니라
# sync_binance_futures_income_v1.py 가 받아온 실제 income API(REALIZED_PNL/COMMISSION/
# FUNDING_FEE) 기록만 신뢰 원천으로 쓴다 — 로컬 CSV는 "주문을 보냈다"만 기록하고
# SL/TP(algo) 가 실제로 언제·얼마에 체결됐는지는 안 남기기 때문.
BINANCE_FUT_INCOME_CSV = os.path.join(BASE_DIR, "coin_binan", "0_binance_usdt_futures_income_v1.csv")

# 성과 tag 표시 순서 + 보유단위 ("day"=스윙 일, "min"=당일 분)
#
# ★ 그룹 분리 (2026-08-16 확정)
#   이 게시판의 목적은 "내 전체 성적표"가 아니라 "봇 태그별 성과 비교"다.
#   3개월 단위로 어느 태그를 늘리고 어느 태그를 죽일지 판단하는 게 목적이므로
#   사람 판단이 섞인 거래는 통합 합산에서 빠져야 한다.
#     bot    = 자동매매. 맨 위 '통합' 행은 이 그룹만 합산한다.
#     alloc  = 자산배분 리밸런싱(통합ETF). 매매전략과 성격이 달라 별도 그룹.
#     manual = 사람이 직접 매수한 태그. 하단 참고행으로만 남긴다.
#   ※ 귀속 기준은 어디까지나 '매수 시점 태그'다. 봇이 산 걸 사람이 팔면 봇 성과,
#     사람이 산 걸 봇이 팔면 수동 성과 (build_round_trips 의 FIFO 흡수 참고).
TAG_ORDER = [
    "통합",
    "주도주", "ROCKET", "ABC-VCP", "TV알림",
    "2X단타", "삼닉v3_저2", "삼닉v3_추세", "삼닉v3_MA", "삼닉v3_미분류",
    "5minHL", "5minHL2", "5minHL_동시", "5minHL_미분류",
    "KR_TR_ORD_B", "KR_TR_ORD_1A0", "KR_TR_VOLUME_1", "KR_TR_VOLUME_2", "KR_TR_VCP1",
    "KR_TR_BASE", "KR_TR_JEO2", "KR_TR_MA", "KR_TR_ADD1",
    "미국VCP", "US_TR_ORD_B", "US_TR_ORD_1A0", "US_TR_VOLUME_1", "US_TR_VOLUME_2",
    "US_TR_VCP1", "US_TR_BASE", "US_TR_JEO2", "US_TR_MA", "US_TR_ADD1",
    "US_TR_LEGACY", "US_TR_LEGACY_TREND", "US_TR_LEGACY_LOW",
    # 옛 이름. 과거 라운드트립이 아직 이 태그라 표에서 사라지지 않게 남겨둔다.
    "KR_TR_VCP2", "US_TR_VCP2",
    "저점사다리",
    # ── 자산배분 (리밸런싱) ──
    "통합ETF",
    # ── 수동 (사람이 직접 매수) ──
    "수동합계", "수동매매", "미국수동", "기타(8042)",
]
TAG_UNIT = {
    "주도주": "day", "ROCKET": "day", "ABC-VCP": "day", "TV알림": "day",
    "수동매매": "day", "통합ETF": "day",
    "미국VCP": "day", "미국수동": "day",
    "KR_TR_ORD_B": "day", "KR_TR_ORD_1A0": "day",
    "KR_TR_VOLUME_1": "day", "KR_TR_VOLUME_2": "day",
    "KR_TR_VCP1": "day", "KR_TR_BASE": "day", "KR_TR_VCP2": "day",
    "KR_TR_JEO2": "day", "KR_TR_MA": "day",
    "KR_TR_ADD1": "day",
    "US_TR_ORD_B": "day", "US_TR_ORD_1A0": "day",
    "US_TR_VOLUME_1": "day", "US_TR_VOLUME_2": "day",
    "US_TR_VCP1": "day", "US_TR_BASE": "day", "US_TR_VCP2": "day",
    "US_TR_JEO2": "day", "US_TR_MA": "day",
    "US_TR_ADD1": "day",
    "US_TR_LEGACY": "day", "US_TR_LEGACY_TREND": "day", "US_TR_LEGACY_LOW": "day",
    "2X단타": "min", "삼닉v3_저2": "min", "삼닉v3_추세": "min", "삼닉v3_MA": "min",
    "삼닉v3_미분류": "min",
    "5minHL": "min", "5minHL2": "min", "5minHL_동시": "min", "5minHL_미분류": "min",
    "기타(8042)": "min", "저점사다리": "min", "통합": "mix", "수동합계": "mix",
}

# 사람이 직접 낸 주문으로 분류되는 태그. 이 태그의 '매도'는 어느 봇 물량을 판 건지
# 알 수 없으므로 자기 서랍이 비면 같은 종목의 봇 서랍에서 FIFO 로 흡수한다
# (build_round_trips 참고). 매수는 그대로 이 태그의 포지션이 된다.
#   ※ 집계에서 빼면 안 된다. 봇이 산 물량을 사람이 판 매도가 흡수될 자리가 없어져
#     그 봇 포지션이 영영 안 닫힌다. "통합 합산·표시에서 제외" 일 뿐이다.
MANUAL_TAGS = {"수동매매", "미국수동", "기타(8042)"}

# 자산배분(주간 리밸런싱). 매매전략이 아니라 비중조절이라 통합에 섞으면 비교가 흐려진다.
ALLOC_TAGS = {"통합ETF"}

# 레거시 편입분(tr_migrate_us_1887.py). 새 장부 이전에 산 물량을 봇이 청산할 수 있게
# 태그만 붙여둔 것이라 '봇 성과'가 아니다. 통합 합산·표시에서 뺀다.
#   ※ MANUAL_TAGS 와 같은 이유로 TAG_ORDER 에서 지우면 안 된다. 집계에서 통째로 빼면
#     이 물량을 판 매도가 FIFO 로 흡수될 자리가 없어져 다른 봇 포지션이 영영 안 닫힌다.
#   ※ 청산이 다 끝나도 되돌리지 말 것. 닫힌 포지션 기록이 장부에 계속 남아 있어서
#     빼는 순간 과거 레거시 손익이 통합에 되살아난다.
LEGACY_TAGS = {"US_TR_LEGACY", "US_TR_LEGACY_TREND", "US_TR_LEGACY_LOW"}

# ───────── 컷오프 시딩분 (2026-09-05 분리) ─────────
# 컷오프(=베이스라인 리셋) 때 실계좌 잔고로 심어놓은 포지션은 '봇이 잡은 자리'가 아니다.
# 진입가가 실제 매수가가 아니라 컷오프 직전 종가로 갈아끼워져 있어서, 이걸 청산한 손익은
# "그 봇이 그 주에 번/잃은 돈"이 아니라 "컷오프 이후 가격변화"일 뿐이다.
# 섞어두면 태그를 살릴지 죽일지 판단이 불가능해진다 — 실제로 2X단타(보유단위 '분')가
# 컷오프 시딩 122630 을 18일에 걸쳐 1~2주씩 흘려판 -151,458원을 뒤집어쓰고 있었다.
#
# 그래서 시딩 서랍만 태그 뒤에 이 접미사를 붙여 별도 행으로 뺀다.
#   · 봇이 컷오프 이후 새로 산 물량은 원래 태그 서랍에 그대로 들어간다(정확 태그 우선).
#   · 시딩 서랍이 남으면 build_round_trips 의 FIFO 흡수가 알아서 가져간다.
#   ※ MANUAL/LEGACY 와 같은 이유로 집계에서 통째로 빼면 안 된다. 이 물량을 판 매도가
#     흡수될 자리가 없어지면 같은 종목의 봇 포지션이 대신 깎여 봇 성과가 오염된다.
#     '통합 합산·봇 카드에서 제외' 일 뿐이다.
SEED_TAG_SUFFIX = "_시딩"


def seed_tag(tag):
    """원래 성과 tag → 컷오프 시딩분 tag."""
    return f"{tag}{SEED_TAG_SUFFIX}"


def is_seed_tag(tag):
    return str(tag).endswith(SEED_TAG_SUFFIX)

# 그룹별 집계 행(개별 태그가 아니라 합산 행). key → (그룹, 합산 대상 판정)
AGGREGATE_TAGS = {"통합": "bot", "수동합계": "manual"}


def tag_group(tag):
    """성과 tag → 그룹. 'bot' | 'alloc' | 'manual' | 'legacy' | 'seed'"""
    if tag in AGGREGATE_TAGS:
        return AGGREGATE_TAGS[tag]
    if is_seed_tag(tag):
        return "seed"
    if tag in MANUAL_TAGS:
        return "manual"
    if tag in ALLOC_TAGS:
        return "alloc"
    if tag in LEGACY_TAGS:
        return "legacy"
    return "bot"


def tag_unit(tag):
    """성과 tag → 보유단위 'day' | 'min' | 'mix'.

    시딩분은 원래 태그가 당일봇('분')이어도 며칠~몇 주에 걸쳐 흘려파는 물량이라
    보유'분'으로 재면 수천 분짜리 숫자만 나온다. 항상 '일'로 잰다.
    """
    if is_seed_tag(tag):
        return "day"
    return TAG_UNIT.get(tag, "min")

# 오래된 이름을 쓰는 보조 출력부와의 호환용 alias.
BOT_ORDER = TAG_ORDER
BOT_UNIT = TAG_UNIT


# ───────── 컷오프 리셋 (make_position_baseline.py 산출물) ─────────
# fill 스트림 replay 만으로는 스트림 시작 이전 보유분·계좌이관을 따라갈 수 없어
# 트래커 잔량이 실제 계좌와 크게 어긋난다. 특정 날짜에 실제 잔고로 한 번 맞추고
# 그 이후부터 깨끗하게 집계한다. 컷오프 이전 주차는 기존 로직 그대로 보존한다.
BASELINE_GLOB = os.path.join(BASE_DIR, "0order", "position_baseline_*.json")


def load_baseline():
    """가장 최근 position_baseline_*.json. 없으면 None(=리셋 없이 기존 동작)."""
    paths = sorted(glob.glob(BASELINE_GLOB))
    if not paths:
        return None
    try:
        with open(paths[-1], encoding="utf-8") as f:
            data = json.load(f)
        data["_path"] = paths[-1]
        return data
    except Exception as e:
        print(f"  ⚠ baseline 읽기 실패({os.path.basename(paths[-1])}): {e}")
        return None


BASELINE = load_baseline()
CUTOFF_DATE = (BASELINE or {}).get("cutoff_date") or None


def _after_cutoff(fill):
    """이 체결이 컷오프 당일 이후인가. 컷오프 미설정이면 항상 False(기존 동작 유지)."""
    return bool(CUTOFF_DATE) and str(fill.get("date", "")) >= CUTOFF_DATE


# ───────── 부분청산 실현손익 (2026-08-16 확정) ─────────
# 기존 집계는 포지션이 완전히 flat 될 때만 손익을 확정했다. 그러면 분할청산이
# 전략의 본질인 태그(KR/US_TR_MA 의 MA5/10/20 1/3 청산, 추세태그의 고점 1/3,
# 주도주의 잔량 보유)가 구조적으로 늦게·뭉쳐서 잡혀 태그별 비교가 왜곡된다.
#
# 결정: 금액은 매도 시점 / 건수는 포지션 완결 기준.
#   · 주간 '실현손익' 금액  → 그 주에 일어난 모든 매도의 실현분 (REALIZATIONS)
#   · 거래건수·승률·PF·기대값 → 완결 라운드트립 기준 (trips) — 기존 그대로
#   분할청산 전략이 승률 계산에서 불리해지지 않게 하려는 것이다
#   (고점 1/3 익절 + 잔량 손절을 1승 2패로 세지 않는다).
#
# 적용 시점은 컷오프(=베이스라인 리셋)부터. 그 이전 주차는 기존 완결 기준을 그대로
# 보존해서 이미 보신 과거 숫자가 흔들리지 않게 한다.
REALIZATIONS = []


def realize_active():
    """부분청산 회계가 켜졌는가. 컷오프가 실제로 도래해야 켠다.

    도래 전에 켜면 TR 이월분 때문에 '아직 오지도 않은 컷오프 주' 행이 미리 생긴다.
    build_round_trips 의 베이스라인 시딩도 같은 조건(도래 후)에서만 도므로 기준을 맞춘다.
    """
    return bool(CUTOFF_DATE) and date.today().isoformat() >= CUTOFF_DATE


def realize_from_week():
    """부분청산 회계를 적용하기 시작하는 주(월요일 iso). 컷오프 전이면 None(=기능 off)."""
    if not realize_active():
        return None
    try:
        return week_monday(datetime.strptime(CUTOFF_DATE, "%Y-%m-%d").date()).isoformat()
    except Exception:
        return None


def _mk_realization(bot, source, code, name, entry_dt, exit_dt, qty, avg_entry, price):
    """매도 1건의 실현손익 레코드. trip 과 같은 스키마라 _krw()/비용계산을 그대로 쓴다."""
    return trade_costs.apply_to_trip({
        "bot": bot, "source": source, "code": code, "name": name,
        "entry_dt": entry_dt, "exit_dt": exit_dt,
        "avg_entry": avg_entry, "avg_exit": price, "qty": qty,
        "pnl_pct": ((price / avg_entry - 1.0) * 100.0) if avg_entry > 0 else 0.0,
        "pnl_amt": (price - avg_entry) * qty,
    })


# ───────── fill 이벤트 수집 (make_danta_journal.py 와 동일 스키마) ─────────
def load_fills():
    """intraday_signals 전체 → 시간순 fill 리스트. 교차파일 중복제거."""
    fills = []
    seen = set()
    for path in sorted(glob.glob(os.path.join(SIGNAL_DIR, "*.json"))):
        m = re.match(r"(\d{4}-\d{2}-\d{2})\.json$", os.path.basename(path))
        if not m:
            continue
        day = m.group(1)
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            print(f"  ⚠ {os.path.basename(path)} 읽기 실패: {e}")
            continue
        for ev in data.get("events", []):
            if ev.get("type") != "fill":
                continue
            side = ev.get("side")
            if side not in ("buy", "sell"):
                continue
            qty   = int(ev.get("qty", 0) or 0)
            price = float(ev.get("price", 0) or 0)
            if qty <= 0 or price <= 0:
                continue
            tm  = str(ev.get("tm", "") or "").strip()
            ono = str(ev.get("ord_no", "") or "").strip()
            # 교차파일 중복제거: (acct,ord_no,side,qty,price,tm,slot) 최초 1건.
            # slot 을 키에 넣는 이유: 8042CHU3 는 한 주문(ord_no)을 slot 별로 나눠
            # 기록한다(예: 30주 매수 = slot"trend" 15주 + slot"ma" 15주, ord_no 동일).
            # slot 이 빠지면 두 줄이 완전히 같아 보여 뒤엣것이 중복으로 버려지고,
            # 그 slot 의 매도는 짝을 못 찾아 통째로 폐기된다(2026-08-03 122630 등 3건).
            if ono:
                dk = (str(ev.get("acct", "")), ono, side, qty, price, tm,
                      str(ev.get("slot") or ""))
                if dk in seen:
                    continue
                seen.add(dk)
            # 체결시각 datetime (당일 보유'분' 계산용). tm(HHMMSS) 우선.
            if len(tm) >= 6 and tm[:6].isdigit():
                hh, mm, ss = int(tm[0:2]), int(tm[2:4]), int(tm[4:6])
            elif len(tm) >= 4 and tm.isdigit():
                hh, mm, ss = int(tm[0:2]), int(tm[2:4]), 0
            else:
                la = str(ev.get("logged_at", "") or "")
                if len(la) >= 5 and la[:2].isdigit():
                    hh, mm = int(la[0:2]), int(la[3:5]); ss = 0
                else:
                    hh, mm, ss = 9, 0, 0
            y, mo, d = int(day[0:4]), int(day[5:7]), int(day[8:10])
            try:
                dt = datetime(y, mo, d, min(hh, 23), min(mm, 59), min(ss, 59))
            except Exception:
                dt = datetime(y, mo, d, 9, 0, 0)
            fills.append({
                "date": day, "dt": dt, "side": side,
                "acct": str(ev.get("acct", "") or ""),
                "strategy": ev.get("strategy"),
                # KR 코드(순수 6자리 숫자)만 zero-pad. US 심볼(EWS/BX 등)은 그대로.
                "code": (lambda c: c.zfill(6) if c.isdigit() else c)(str(ev.get("code", ""))),
                "name": str(ev.get("name", "") or ""),
                "signal": ev.get("signal"),
                "slot": ev.get("slot"),
                "qty": qty, "price": price, "ord_no": ono,
            })
    # *DIP ↔ 원계좌 교차기록 중복제거: 같은 날 같은 ord_no 가 1887DIP/8042DIP 로도 잡혀있으면
    # sync_1887_fills.py/sync_8042_fills.py 가 태깅 없이 재생성한 acct="1887"/"8042" 사본을
    # 버린다(*DIP 쪽이 정확한 전략태그를 갖는 원본이므로 그쪽만 남긴다).
    # 8042CHU3(삼닉v3) 도 같은 구조다. 봇은 slot 을 실어 acct="8042CHU3" 로 기록하는데,
    # 장 마감 후 동기화가 같은 ord_no 를 slot 없이 acct="8042CHU"/"8042" 로 재기록한다
    # (origin="manual", logged_at 22:53 대). slot 있는 8042CHU3 쪽이 원본이므로 그쪽만 남긴다.
    for src_acct, dup_acct in (("1887DIP", "1887"), ("8042DIP", "8042"),
                               ("8042CHU3", "8042CHU"), ("8042CHU3", "8042"),
                               ("8042DIP", "8042CHU")):
        src_ordnos = {(f["date"], f["ord_no"]) for f in fills if f["acct"] == src_acct and f["ord_no"]}
        if src_ordnos:
            fills = [f for f in fills
                     if not (f["acct"] == dup_acct and f["ord_no"] and (f["date"], f["ord_no"]) in src_ordnos)]
    fills.sort(key=lambda x: x["dt"])
    return fills


def resolve_performance_tag(fill):
    """매수 fill 기준 성과 tag. 매도 signal은 성과 분류 기준으로 쓰지 않는다."""
    acct = str(fill.get("acct", "") or "")
    strategy = fill.get("strategy")
    slot = str(fill.get("slot", "") or "").lower()
    signal = str(fill.get("signal", "") or "").lower()
    if acct in ("1887DIP", "8042DIP"):
        return "저점사다리"
    s = (strategy or "").lower()
    if acct == "1887":
        if s == "rocket":
            return "ROCKET"
        if s == "vcp":
            return "ABC-VCP"
        if s == "leader":
            return "주도주"
        if s == "tvalert":
            return "TV알림"                     # 자체청산 봇. 사람 주문 아님
        return "수동매매"                       # manual/미태깅 → 사람이 직접 낸 주문
    if acct == "1887US":
        if s == "usvcp":
            return "미국VCP"
        if s in ("us_pine_buy", "us_pine_lowbuy"):
            return None                         # Pine TR은 ledger strategy_tag 기준으로 별도 집계
        return "미국수동"
    if acct == "8042ETF":
        return "통합ETF"                        # allone 8042 통합ETF (국내+미국상장 ETF)
    if acct == "8042CHU":
        return "2X단타"
    if acct == "8042CHU3" and s == "trend3":
        if slot == "jeo2":
            return "삼닉v3_저2"
        if slot == "trend":
            return "삼닉v3_추세"
        if slot == "ma":
            return "삼닉v3_MA"
        return "삼닉v3_미분류"
    if acct == "2773":
        if signal == "reconcile":
            return "수동매매"            # reconcile_2773.py 가 남긴 '사람이 MTS에서 청산' 보정분
        if signal == "jeo":
            return "5minHL"
        if signal == "jeo2":
            return "5minHL2"
        if signal == "jeo_both":
            return "5minHL_동시"        # 저·저2 동시 발생. 임의 귀속하지 않고 분리 집계
        if s:
            return "5minHL_미분류"       # 봇인데 signal 만 빠진 경우
        # 컷오프 이후: 봇은 반드시 strategy 를 실어 기록한다(0_low_high_5min_danta.py).
        # 따라서 strategy 가 비어있으면 사람이 직접 낸 주문이다.
        # 컷오프 이전 체결은 봇도 strategy 를 안 실었으므로 기존 분류를 유지한다.
        return "수동매매" if _after_cutoff(fill) else "5minHL_미분류"
    if acct == "8042":
        # 본계좌. lowhigh(실험용)는 build_round_trips 에서 별도 제외된다.
        if s and s != "lowhigh":
            return "기타(8042)"
        return "수동매매" if _after_cutoff(fill) else "기타(8042)"
    return "기타(8042)"


def resolve_owner(acct, strategy):
    """이전 함수명 호환용. 새 집계는 resolve_performance_tag(fill)를 사용한다."""
    return resolve_performance_tag({"acct": acct, "strategy": strategy})


# ───────── 포지션 lifecycle → 완결 라운드트립 ─────────
def build_round_trips(fills, baseline=None, verbose=True):
    """계좌(silo)+종목+성과tag별 평균단가 lifecycle 재구성.
    flat→매수…→flat 한 사이클 = 1 라운드트립(완결거래).
    매수가 성과tag를 확정하고, 매도는 그 포지션에 흡수한다.

    컷오프(baseline) 가 있으면 그 날짜 첫 체결 직전에 열린 포지션을 전부 버리고
    실제 계좌잔고로 다시 시딩한다. 컷오프 이전 주차는 기존 동작 그대로 남는다.

    수동매도 흡수(컷오프 이후에만):
      사람이 MTS 에서 판 매도는 어느 봇 물량인지 알 수 없다. 자기 서랍(수동매매)이
      부족하면 같은 (계좌,종목)의 다른 봇 서랍을 '먼저 산 것부터' 흡수한다.
      → 봇이 사고 사람이 판 거래도 매수 소유봇 성과로 귀속된다.
      귀속 순서(FIFO)는 규칙이지 사실이 아니다. 증권사는 어느 물량을 팔았는지
      기록하지 않으므로, LIFO 로 바꾸면 봇별 성과가 달라진다.
    """
    if baseline is None:
        baseline = BASELINE
    open_pos = {}      # (acct, code, performance_tag) → dict
    trips = []
    REALIZATIONS.clear()   # 재실행(full/web 2회 호출) 시 중복 누적 방지
    cutoff = (baseline or {}).get("cutoff_date")
    realize_on = bool(cutoff) and realize_active()
    seeded = False
    stats = {"seeded": 0, "absorbed": 0, "absorbed_qty": 0,
             "dropped": 0, "dropped_qty": 0}

    def new_pos(owner, name):
        return {"owner": owner, "name": name,
                "buy_qty": 0, "buy_cost": 0.0,
                "sell_qty": 0, "sell_proceeds": 0.0,
                "first_buy_dt": None, "last_sell_dt": None,
                "held": 0}  # 현재 보유수량

    def close_trip(code, p, fallback_dt):
        """held<=0 인 포지션 → 라운드트립 확정."""
        if p["buy_qty"] <= 0 or p["sell_qty"] <= 0:
            return
        avg_entry = p["buy_cost"] / p["buy_qty"]
        avg_exit  = p["sell_proceeds"] / p["sell_qty"]
        if avg_entry <= 0:
            return
        entry_dt = p["first_buy_dt"]
        exit_dt  = p["last_sell_dt"] or fallback_dt
        trips.append(trade_costs.apply_to_trip({
            "bot": p["owner"], "source": "intraday", "code": code, "name": p["name"],
            "entry_dt": entry_dt, "exit_dt": exit_dt,
            "avg_entry": avg_entry, "avg_exit": avg_exit,
            "qty": p["sell_qty"],
            "pnl_pct": (avg_exit / avg_entry - 1.0) * 100.0,
            "pnl_amt": p["sell_proceeds"] - avg_entry * p["sell_qty"],
            "hold_min": max(0.0, (exit_dt - entry_dt).total_seconds() / 60.0),
            "hold_day": max(0, (exit_dt.date() - entry_dt.date()).days),
        }))

    def seed_from_baseline():
        """컷오프 시점: 그때까지의 열린 포지션(유령 포함) 전량 폐기 후 실제 잔고로 재시딩."""
        open_pos.clear()
        try:
            entry_dt = datetime.strptime(cutoff, "%Y-%m-%d").replace(hour=9)
        except Exception:
            entry_dt = datetime.now()
        for b in (baseline.get("positions") or []):
            qty = int(b.get("qty") or 0)
            px  = float(b.get("entry_price") or 0)
            if qty <= 0 or px <= 0:
                continue
            # 시딩분은 원래 태그와 다른 서랍에 심는다(SEED_TAG_SUFFIX 주석 참고).
            # 컷오프 이후 봇이 새로 사면 원래 태그 서랍이 따로 생기고, 매도는
            # '정확 태그 우선 → FIFO' 규칙에 따라 봇 물량부터 소진된다.
            stag = seed_tag(b["tag"])
            key = (b["acct"], b["code"], stag)
            p = open_pos.get(key)
            if p is None:
                p = new_pos(stag, b.get("name", ""))
                p["first_buy_dt"] = entry_dt
                open_pos[key] = p
            p["buy_qty"]  += qty
            p["buy_cost"] += qty * px
            p["held"]     += qty
            stats["seeded"] += 1

    def consume(key, p, want, price, dt):
        """포지션 p 에서 want 만큼 매도 흡수. 완결되면 trip 확정하고 서랍을 비운다."""
        take = min(want, p["held"])
        if take <= 0:
            return 0
        # 부분청산 실현손익: 매도할 때마다 그 시점 평균단가로 기록한다.
        # 컷오프 이전 매도는 남기지 않는다 — 어차피 seed_from_baseline() 이 그때 열려있던
        # 포지션을 전량 폐기하고 실계좌 잔고로 다시 시딩하므로 원가 기준이 갈아엎어진다.
        if realize_on and p["buy_qty"] > 0 and dt.date().isoformat() >= cutoff:
            REALIZATIONS.append(_mk_realization(
                p["owner"], "intraday", key[1], p["name"],
                p["first_buy_dt"] or dt, dt, take,
                p["buy_cost"] / p["buy_qty"], price))
        p["sell_qty"]      += take
        p["sell_proceeds"] += take * price
        p["held"]          -= take
        p["last_sell_dt"]   = dt
        if p["held"] <= 0:
            close_trip(key[1], p, dt)
            open_pos.pop(key, None)
        return take

    for f in fills:
        if cutoff and not seeded and f["date"] >= cutoff:
            seed_from_baseline()
            seeded = True
        perf_tag = resolve_performance_tag(f)
        if perf_tag is None:
            continue
        # 8042 lowhigh = 실험용(며칠 전 2773 으로 이관). 통합 포함 통계에서 완전 제외.
        if f["acct"] == "8042" and (f["strategy"] or "").lower() == "lowhigh":
            continue
        key = (f["acct"], f["code"], perf_tag)
        if f["side"] == "buy":
            p = open_pos.get(key)
            if p is None or p["held"] <= 0:
                # 새 포지션 시작 — 매수 시점 성과tag 확정
                p = new_pos(perf_tag, f["name"])
                open_pos[key] = p
                p["first_buy_dt"] = f["dt"]
            if p["first_buy_dt"] is None:
                p["first_buy_dt"] = f["dt"]
            if f["name"]:
                p["name"] = f["name"]
            p["buy_qty"]  += f["qty"]
            p["buy_cost"] += f["qty"] * f["price"]
            p["held"]     += f["qty"]
        else:  # sell
            need = f["qty"]
            p = open_pos.get(key)
            if p is not None and p["held"] > 0:
                need -= consume(key, p, need, f["price"], f["dt"])
            # 자기 태그 서랍으로 다 못 채운 매도는 같은 (계좌,종목)의 다른 서랍에서 흡수한다.
            # 매도 태그를 신뢰할 수 없는 경로가 여럿이기 때문:
            #   · 사람이 MTS 에서 판 매도 → 어느 봇 물량인지 정보 자체가 없음
            #   · 2773 봇은 매수에만 signal 을 싣고 매도엔 안 실음(place_order 호출부)
            #   · rocket 이 산 걸 출근 자동매도(leader 태그)가 파는 등 봇 간 교차청산
            # 정확한 태그 일치를 먼저 쓰므로, 삼닉v3 처럼 한 종목을 MA/추세 두 슬롯으로
            # 동시 보유하며 각자 자기 태그로 파는 경우의 분리 집계는 그대로 유지된다.
            if need > 0 and _after_cutoff(f):
                cands = sorted(
                    [(v["first_buy_dt"] or datetime.max, k) for k, v in open_pos.items()
                     if v["held"] > 0 and k[0] == f["acct"] and k[1] == f["code"] and k != key],
                    key=lambda x: x[0])
                for _, k2 in cands:
                    if need <= 0:
                        break
                    p2 = open_pos.get(k2)
                    if p2 is None or p2["held"] <= 0:
                        continue
                    got = consume(k2, p2, need, f["price"], f["dt"])
                    if got > 0:
                        need -= got
                        stats["absorbed"] += 1
                        stats["absorbed_qty"] += got
            if need > 0:
                # 대응 매수를 끝내 못 찾은 매도. 예전엔 조용히 사라졌다 → 카운트해서 드러낸다.
                stats["dropped"] += 1
                stats["dropped_qty"] += need

    # 컷오프 날짜는 지났는데 그 이후 체결이 하나도 없으면(봇 휴장·무거래) 위 루프에서
    # 시딩이 안 걸린다. 그대로 두면 컷오프 이전 유령이 계속 남으므로 여기서 시딩한다.
    if cutoff and not seeded and date.today().isoformat() >= cutoff:
        seed_from_baseline()
        seeded = True

    # 미청산 포지션을 남겨 둔다(check_position_drift 가 실제 잔고와 대조).
    global LAST_OPEN_POS
    LAST_OPEN_POS = {k: v for k, v in open_pos.items() if v["held"] > 0}

    if verbose:
        if seeded:
            where = os.path.basename((baseline or {}).get("_path", ""))
            print(f"  컷오프 {cutoff} — 베이스라인 {stats['seeded']}종목 시딩 ({where})")
        elif cutoff:
            print(f"  컷오프 {cutoff} 예정 — 아직 도래 전이라 기존 집계 유지")
        if stats["absorbed"]:
            print(f"  수동매도 흡수: {stats['absorbed']}건 {stats['absorbed_qty']}주 (봇 서랍에서 FIFO)")
        if stats["dropped"]:
            print(f"  ⚠ 짝 못 찾은 매도: {stats['dropped']}건 {stats['dropped_qty']}주 "
                  f"(집계 제외 — 컷오프 이전 보유분이면 정상)")
    return trips


# 마지막 build_round_trips 의 미청산 포지션. check_position_drift 가 읽는다.
LAST_OPEN_POS = {}

# 게시판 계좌라벨 → 실제 증권계좌. 드리프트 대조용.
SILO_TO_ACCOUNT = {
    "1887": "1887", "1887US": "1887", "1887DIP": "1887",
    "8042": "8042", "8042ETF": "8042", "8042CHU": "8042",
    "8042CHU3": "8042", "8042DIP": "8042",
    "2773": "2773",
}


def check_position_drift(verbose=True):
    """트래커 미청산 + TR장부 미청산  vs  실제 계좌잔고 대조.

    둘이 어긋나면 집계가 조용히 틀어진다. 실제로 컷오프 전까지
    8042 228790 이 트래커 1953주 vs 실제 225주까지 벌어져 있었다.
    주 원인은 fill 로 안 잡히는 경로다 — 계좌 간 대체입출고(7/31 0193T0/0193W0
    1887→8042 이관), 사람이 MTS 에서 낸 청산, 스트림 시작 이전 보유분.
    소급 복구는 불가능하므로 '드러내는' 것이 목적이다. 크게 벌어지면
    make_position_baseline.py 로 컷오프를 다시 잡는다.

    반환: [(계좌, 종목, 트래커수량, 실제수량, 차이), ...]  (차이 있는 것만)
    """
    tracked = defaultdict(int)
    for (silo, code, _tag), p in LAST_OPEN_POS.items():
        acct = SILO_TO_ACCOUNT.get(silo, silo)
        tracked[(acct, str(code).upper())] += int(p["held"])

    # TR 자체장부 미청산분도 실제 보유이므로 합산한다.
    for filename, acct in (("tr_ledger_2773_KR.json", "2773"),
                           ("tr_ledger_1887_US.json", "1887")):
        path = os.path.join(TR_LEDGER_DIR, filename)
        if not os.path.exists(path):
            continue
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue
        for pos in (data.get("positions") or {}).values():
            rem = int(pos.get("remaining_strategy_qty", 0) or 0)
            if rem > 0:
                tracked[(acct, str(pos.get("code", "")).upper())] += rem

    actual = defaultdict(int)
    have_balance = False
    for fname, acct in (("holdings_1887.json", "1887"),
                        ("holdings_8042.json", "8042"),
                        ("holdings_2773.json", "2773")):
        path = os.path.join(SCRIPT_DIR, fname)
        if not os.path.exists(path):
            continue
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue
        have_balance = True
        for h in (data.get("holdings") or []):
            code = str(h.get("stock_code") or "").upper()
            if code:
                actual[(acct, code)] += int(h.get("quantity") or 0)
        for h in (data.get("us_holdings") or []):
            code = str(h.get("ticker") or "").upper()
            if code and code != "NONE":
                actual[(acct, code)] += int(h.get("quantity") or 0)

    if not have_balance:
        # 잔고파일이 아예 없으면 '드리프트 0' 과 구분되어야 한다(배너를 띄우면 안 됨).
        return None

    diffs = []
    for k in sorted(set(tracked) | set(actual)):
        t, a = tracked.get(k, 0), actual.get(k, 0)
        if t != a:
            diffs.append((k[0], k[1], t, a, t - a))

    if verbose and diffs:
        print(f"\n  ⚠ 잔고 드리프트 {len(diffs)}건 (트래커+TR장부 vs 실제계좌)")
        for acct, code, t, a, d in diffs[:20]:
            print(f"     {acct:5s} {code:8s} 트래커 {t:>6}주  실제 {a:>6}주  차이 {d:>+6}")
        if len(diffs) > 20:
            print(f"     … 외 {len(diffs)-20}건")
        print("     크게 벌어졌으면: python 0order/make_position_baseline.py --apply 로 컷오프 재설정")
    return diffs


def build_drift_payload(diffs, limit=15):
    """check_position_drift() 결과 → 웹 게시판 배너용 dict.

    이 경고는 지금까지 콘솔에만 찍혀서, 배치를 돌린 사람이 로그를 안 보면
    '트래커엔 있는데 계좌엔 0주' 인 유령 포지션이 몇 주씩 쌓여도 몰랐다.
    유령은 라운드트립이 영영 안 닫혀서 그 태그 성과가 통째로 안 잡힌다
    (미국VCP 가 계속 '0건' 으로 보이던 이유). 표 위에 띄워서 드러낸다.

    diffs 가 None(잔고파일 없음) 이면 None — 배너를 그리지 않는다.
    """
    if diffs is None:
        return None
    rows = []
    for acct, code, tracked, actual, diff in diffs:
        rows.append({"acct": acct, "code": code,
                     "tracked": tracked, "actual": actual, "diff": diff,
                     "ghost": bool(tracked > 0 and actual == 0)})
    # 유령 먼저, 그 안에서는 수량차 큰 순
    rows.sort(key=lambda r: (not r["ghost"], -abs(r["diff"])))
    return {
        "count": len(rows),
        "ghost": sum(1 for r in rows if r["ghost"]),
        "cutoff": CUTOFF_DATE,
        "checked_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "rows": rows[:limit],
        "more": max(0, len(rows) - limit),
        "fix_hint": "python 0order/make_position_baseline.py --apply",
    }


def _parse_dt(value, fallback_day=None):
    text = str(value or "").strip()
    if text:
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
            pass
    if fallback_day:
        try:
            return datetime.strptime(str(fallback_day)[:10], "%Y-%m-%d")
        except Exception:
            pass
    return None


def _tr_position_realizations(pos, tag, cutoff_date, carry_out):
    """TR ledger 포지션 1개를 주문 시간순으로 replay → 매도 건별 실현손익.

    ledger 는 완전청산 포지션만 성과로 내주는데(remaining_strategy_qty>0 이면 skip),
    MA 1/3 청산·고점 1/3 익절이 바로 그 '완전청산 전' 상태라 성과가 통째로 안 잡힌다.
    여기서 부분매도마다 그 시점 평균단가로 실현손익을 뽑는다.

    컷오프 이전 매도는 두 갈래다.
      · 컷오프 전에 완전청산된 포지션 → 그 주 trip 으로 이미 집계됨 → 버린다.
      · 컷오프 시점에 아직 열려 있던 포지션 → 완결 기준으로도 집계된 적이 없고,
        새 회계에서는 전량청산 시점에도 안 잡힌다 → carry_out 으로 넘겨 컷오프 주에 이월.
        (intraday 쪽은 baseline 재시딩이 원가를 갈아엎으므로 이월하지 않는다)
    """
    evs = []
    for o in (pos.get("orders") or []):
        q = int(o.get("filled_qty", 0) or 0)
        px = float(o.get("avg_fill_price") or o.get("order_price") or 0)
        dt = _parse_dt(o.get("ordered_at"), o.get("trade_date"))
        if q <= 0 or px <= 0 or dt is None:
            continue
        evs.append((dt, o.get("side"), q, px))
    evs.sort(key=lambda x: x[0])

    code = pos.get("code", "")
    name = pos.get("name", "")
    held, cost, first_buy = 0, 0.0, None
    pending, out, crossed = [], [], False
    for dt, side, q, px in evs:
        if not crossed and dt.date().isoformat() >= cutoff_date:
            crossed = True
            if held > 0:
                carry_out.extend(pending)
            pending = []
        if side == "BUY":
            if held <= 0:
                first_buy = dt
            cost += q * px
            held += q
        elif side == "SELL":
            take = min(q, held)
            if take <= 0:
                continue
            avg = cost / held
            r = _mk_realization(tag, "tr_ledger", code, name,
                                first_buy or dt, dt, take, avg, px)
            cost -= take * avg
            held -= take
            if crossed:
                out.append(r)
            else:
                pending.append(r)
                if held <= 0:
                    pending = []      # 컷오프 전 완결 → trip 으로 이미 집계
    if not crossed and held > 0:
        carry_out.extend(pending)
    return out


def build_tr_ledger_round_trips():
    """Pine TR ledger의 code|strategy_tag 포지션에서 완전 청산된 라운드트립을 만든다.
    같은 replay 로 부분매도 실현손익(REALIZATIONS)도 함께 채운다."""
    trips = []
    cutoff = CUTOFF_DATE if realize_active() else None
    carry = []
    for filename in ("tr_ledger_2773_KR.json", "tr_ledger_1887_US.json"):
        path = os.path.join(TR_LEDGER_DIR, filename)
        if not os.path.exists(path):
            continue
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            print(f"  ⚠️ TR ledger 읽기 실패({filename}): {e}")
            continue

        positions = data.get("positions") or {}
        for pos in positions.values():
            tag = pos.get("strategy_tag")
            if not tag:
                continue
            # 실현손익 replay 는 미청산 포지션도 대상 — 아래 완결 트립 로직보다 먼저 돈다.
            if cutoff:
                REALIZATIONS.extend(_tr_position_realizations(pos, tag, cutoff, carry))
            orders = pos.get("orders") or []
            buys = [o for o in orders if o.get("side") == "BUY" and int(o.get("filled_qty", 0) or 0) > 0]
            sells = [o for o in orders if o.get("side") == "SELL" and int(o.get("filled_qty", 0) or 0) > 0]
            buy_qty = sum(int(o.get("filled_qty", 0) or 0) for o in buys)
            sell_qty = sum(int(o.get("filled_qty", 0) or 0) for o in sells)
            if buy_qty <= 0 or sell_qty <= 0:
                continue
            if int(pos.get("remaining_strategy_qty", max(buy_qty - sell_qty, 0)) or 0) > 0:
                continue

            buy_cost = sum(int(o.get("filled_qty", 0) or 0) * float(o.get("avg_fill_price") or o.get("order_price") or 0) for o in buys)
            sell_proceeds = sum(int(o.get("filled_qty", 0) or 0) * float(o.get("avg_fill_price") or o.get("order_price") or 0) for o in sells)
            if buy_cost <= 0 or sell_proceeds <= 0:
                continue
            avg_entry = buy_cost / buy_qty
            avg_exit = sell_proceeds / sell_qty
            entry_dt = min((_parse_dt(o.get("ordered_at"), o.get("trade_date")) for o in buys), default=None)
            exit_dt = max((_parse_dt(o.get("ordered_at"), o.get("trade_date")) for o in sells), default=None)
            if entry_dt is None or exit_dt is None:
                continue
            closed_qty = min(buy_qty, sell_qty)
            hold_min = max(0.0, (exit_dt - entry_dt).total_seconds() / 60.0)
            hold_day = max(0, (exit_dt.date() - entry_dt.date()).days)
            trips.append(trade_costs.apply_to_trip({
                "bot": tag, "source": "tr_ledger", "code": pos.get("code", ""), "name": pos.get("name", ""),
                "entry_dt": entry_dt, "exit_dt": exit_dt,
                "avg_entry": avg_entry, "avg_exit": avg_exit,
                "qty": closed_qty,
                "pnl_pct": (avg_exit / avg_entry - 1.0) * 100.0,
                "pnl_amt": (avg_exit - avg_entry) * closed_qty,
                "hold_min": hold_min, "hold_day": hold_day,
            }))

    # 컷오프 시점 미청산 포지션의 이전 부분매도 → 컷오프 주로 이월(총손익 보존).
    if carry and cutoff:
        try:
            cut_dt = datetime.strptime(cutoff, "%Y-%m-%d").replace(hour=9)
        except Exception:
            cut_dt = datetime.now()
        for r in carry:
            r["carry_in"] = r["exit_dt"].date().isoformat()
            r["exit_dt"] = cut_dt
        REALIZATIONS.extend(carry)
        amt = sum(_krw(r, "net_pnl_amt") for r in carry)
        print(f"  컷오프 이월(TR 미청산 포지션의 이전 부분매도): {len(carry)}건 {amt:,.0f}원")
    return trips


# ───────── 주(월~금) 버킷 ─────────
def week_monday(d):
    """date d 가 속한 주의 월요일(date)."""
    return d - timedelta(days=d.weekday())


def week_label(monday):
    fri = monday + timedelta(days=4)
    return f"{monday.month:02d}/{monday.day:02d}~{fri.month:02d}/{fri.day:02d}"


_WDAY_KO = "월화수목금토일"


def day_label(day_iso):
    """'2026-08-17' → '08/17(월)'."""
    try:
        d = datetime.strptime(day_iso, "%Y-%m-%d").date()
    except Exception:
        return day_iso
    return f"{d.month:02d}/{d.day:02d}({_WDAY_KO[d.weekday()]})"


def _krw(t, field):
    """체결통화 금액 → 원화. 미국 체결(USD)만 환산.
    환율은 '청산 주차'에 고정된 값(fx_for_week) — 과거 주차는 재실행해도 동일."""
    v = float(t.get(field) or 0.0)
    if t.get("ccy") != "USD":
        return v
    return v * fx_for_week(week_monday(t["exit_dt"].date()))


def summarize(trips, unit, reals=None):
    """라운드트립 리스트 → 그림4-2 지표 dict. unit: 'day'|'min'|'mix'.
    승률·평균·손익비·기대값·PF 는 전부 수수료·세금 차감 후(net) 기준이다.

    reals 가 주어지면(컷오프 이후 주차) 실현손익 '금액' 3개 컬럼만 매도 건별 실현분으로
    바꾼다. 거래건수·승률·PF·기대값은 그대로 완결 라운드트립 기준이다.
    → 분할청산 전략이 승률에서 불리해지지 않으면서 금액은 제때 반영된다.
    """
    if reals is None:
        amt_src, sells = trips, None
    else:
        amt_src, sells = reals, len(reals)
    sums = {
        "sum_pnl_amt":   int(round(sum(_krw(t, "net_pnl_amt") for t in amt_src))),
        "sum_pnl_gross": int(round(sum(_krw(t, "pnl_amt") for t in amt_src))),
        "sum_fee_tax":   int(round(sum(_krw(t, "fee_tax") for t in amt_src))),
        "sell_events":   sells,
    }
    n = len(trips)
    if n == 0:
        # 완결 포지션은 없어도 부분매도 실현손익은 있을 수 있다(잔량 보유 중).
        empty = {"trades": 0, "winrate": None, "pl_ratio": None,
                 "avg_win": None, "avg_loss": None,
                 "max_win": None, "max_loss": None, "hold_win": None, "hold_loss": None,
                 "expectancy": None, "expectancy_amt": None, "profit_factor": None}
        empty.update(sums)
        return empty
    wins   = [t for t in trips if t["net_pnl_pct"] > 0]
    losses = [t for t in trips if t["net_pnl_pct"] <= 0]

    def avg(xs):
        return round(sum(xs) / len(xs), 2) if xs else None

    # 손익비 = 평균수익금액 / 평균손실금액 (원화 환산, 기준점 1.0)
    win_amts  = [_krw(t, "net_pnl_amt") for t in wins]
    loss_amts = [abs(_krw(t, "net_pnl_amt")) for t in losses]
    avg_win_amt  = (sum(win_amts) / len(win_amts)) if win_amts else 0.0
    avg_loss_amt = (sum(loss_amts) / len(loss_amts)) if loss_amts else 0.0
    pl_ratio = round(avg_win_amt / avg_loss_amt, 2) if avg_loss_amt > 0 else None

    # Profit Factor = 총이익 / 총손실. 1.0 이 손익분기, 1.5 이상이면 견고.
    gross_profit = sum(win_amts)
    gross_loss   = sum(loss_amts)
    profit_factor = round(gross_profit / gross_loss, 2) if gross_loss > 0 else None

    # Expectancy = 1거래당 기대 수익. %(통화무관)와 금액(원화) 둘 다 낸다.
    expectancy     = round(sum(t["net_pnl_pct"] for t in trips) / n, 3)
    expectancy_amt = int(round(sum(_krw(t, "net_pnl_amt") for t in trips) / n))

    # 통합(mix)은 일/분이 섞여 보유시간 평균이 무의미 → 보유 컬럼 공백
    if unit == "mix":
        hold_win = hold_loss = None
    else:
        hk = "hold_day" if unit == "day" else "hold_min"
        hold_win  = avg([t[hk] for t in wins])
        hold_loss = avg([t[hk] for t in losses])

    out = {
        "trades": n,
        "winrate": round(len(wins) / n * 100.0, 1),
        "pl_ratio": pl_ratio,
        "avg_win":  avg([t["net_pnl_pct"] for t in wins]),
        "avg_loss": avg([abs(t["net_pnl_pct"]) for t in losses]),
        "max_win":  round(max([t["net_pnl_pct"] for t in wins]), 2) if wins else None,
        "max_loss": round(max([abs(t["net_pnl_pct"]) for t in losses]), 2) if losses else None,
        "hold_win":  hold_win,
        "hold_loss": hold_loss,
        "expectancy": expectancy,
        "expectancy_amt": expectancy_amt,
        "profit_factor": profit_factor,
    }
    out.update(sums)
    return out


def build_report(trips, weeks_limit=None):
    """weeks_limit 지정 시 최근 N주만 포함(총계도 그 구간 기준 재계산).
    None 이면 전체 누적."""
    reals = list(REALIZATIONS)
    rf = realize_from_week()          # 이 주차부터 금액을 매도시점 실현으로 집계

    def _wk(x):
        return week_monday(x["exit_dt"].date())

    all_weeks = sorted({_wk(t) for t in trips} |
                       {_wk(r) for r in reals if rf and _wk(r).isoformat() >= rf})
    if weeks_limit:
        latest = max(all_weeks[-1] if all_weeks else week_monday(date.today()), week_monday(date.today()))
        keep = {latest - timedelta(days=7 * i) for i in range(weeks_limit)}
        trips = [t for t in trips if _wk(t) in keep]
        reals = [r for r in reals if _wk(r) in keep]
        all_weeks = sorted(keep)
    elif not all_weeks:
        all_weeks = [week_monday(date.today())]

    # 태그별 → 주별 그룹
    by_bot = defaultdict(list)
    for t in trips:
        by_bot[t["bot"]].append(t)
    by_real = defaultdict(list)
    for r in reals:
        by_real[r["bot"]].append(r)

    # 합산 행. '통합' 은 봇 그룹만 — 수동매매·통합ETF 는 섞지 않는다.
    for agg, grp in AGGREGATE_TAGS.items():
        by_bot[agg] = [t for t in trips if tag_group(t["bot"]) == grp]
        by_real[agg] = [r for r in reals if tag_group(r["bot"]) == grp]

    week_keys = [w.isoformat() for w in all_weeks]
    week_labels = {w.isoformat(): week_label(w) for w in all_weeks}

    ordered_tags = list(BOT_ORDER)
    for extra in sorted(k for k in set(by_bot) | set(by_real) if k not in ordered_tags):
        ordered_tags.append(extra)

    def _summ(tag, unit, trip_list, real_list, use_reals):
        return summarize(trip_list, unit, real_list if use_reals else None)

    bots_out = []
    for bot in ordered_tags:
        bt = by_bot.get(bot, [])
        br = by_real.get(bot, [])
        unit = tag_unit(bot)
        weeks = []
        for w in all_weeks:
            wk = w.isoformat()
            use_reals = bool(rf) and wk >= rf
            wt = [t for t in bt if _wk(t).isoformat() == wk]
            wr = [r for r in br if _wk(r).isoformat() == wk]
            s = _summ(bot, unit, wt, wr, use_reals)
            s["week"] = wk
            s["label"] = week_labels[wk]
            s["fx_usdkrw"] = round(fx_for_week(w), 2)   # 그 주 원화 환산에 쓴 환율(복기용)
            weeks.append(s)
        # TOTAL 은 주차 합이라 컷오프 전/후가 섞인다 → 주차별로 고른 금액을 그대로 더한다.
        total = summarize(bt, unit)
        for k in ("sum_pnl_amt", "sum_pnl_gross", "sum_fee_tax"):
            total[k] = sum(w[k] for w in weeks)
        total["sell_events"] = sum((w.get("sell_events") or 0) for w in weeks) or None
        total["week"] = "TOTAL"
        total["label"] = (f"최근 {weeks_limit}주" if weeks_limit else "전체")
        bots_out.append({
            "key": bot,
            "group": tag_group(bot),
            "unit": ("일" if unit == "day" else ("분" if unit == "min" else "혼합")),
            "weeks": weeks,
            "total": total,
        })

    return {
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "weeks_limit": weeks_limit,
        "week_keys": week_keys,
        "week_labels": week_labels,
        # 부분청산 회계가 켜지는 주(월요일). 이 주부터 실현손익 금액은 매도시점 기준.
        "realize_from": rf,
        "cutoff_date": CUTOFF_DATE,
        # 주차별 환산환율(=weekly_fx_snapshot.json 과 동일). 이 파일만 있어도 재현 가능.
        "fx_usdkrw": {w.isoformat(): round(fx_for_week(w), 2) for w in all_weeks},
        "bots": bots_out,
        "quarter": build_quarter(by_bot, by_real, all_weeks, ordered_tags, rf),
    }


# ───────── 일별 버킷 (당일 성과 확인용) ─────────
# 주간표는 월요일부터 뭉쳐서 "오늘 뭘 했나"가 안 보인다(월요일에만 우연히 보인다).
# 라운드트립·실현손익 레코드는 둘 다 exit_dt 를 들고 있으므로 집계 규칙은 그대로 두고
# 버킷만 주 → 날짜로 바꾼다. 새로 계산하는 값은 없다.
def build_daily_report(trips, days_limit=None):
    """일자별 태그 성과. build_report 의 주차 그룹핑을 날짜로 바꾼 판.

    주간표와 숫자가 어긋나면 안 되므로 규칙을 그대로 따른다.
      · 거래건수·승률·PF·기대값 = 완결 라운드트립(trips) 기준
      · 컷오프 이후 '금액'      = 그 날 매도의 실현분(REALIZATIONS) 기준
    realize_from_week() 가 주 월요일 iso 라서 날짜 문자열 비교로도 같은 구간이 잘린다.
    → 한 주에 속한 일별 금액을 더하면 그 주의 주간 금액과 정확히 일치한다.

    활동(완결거래·매도실현) 있는 날짜·태그만 담는다. 전 태그 × 전 거래일을 다 채우면
    JSON 이 몇 배로 불어나는데 대부분이 0 행이라 쓸모가 없다.
    days_limit 지정 시 활동 있는 날짜 중 최근 N 일만.
    """
    reals = list(REALIZATIONS)
    rf = realize_from_week()

    def _dk(x):
        return x["exit_dt"].date().isoformat()

    all_days = sorted({_dk(t) for t in trips} |
                      {_dk(r) for r in reals if rf and _dk(r) >= rf})
    if days_limit:
        all_days = all_days[-days_limit:]
    keep = set(all_days)

    # (태그, 날짜) 로 미리 쪼개둔다. 날짜마다 전체 리스트를 훑으면 O(일×태그×건수).
    by_bot_day = defaultdict(list)
    by_real_day = defaultdict(list)
    for t in trips:
        dk = _dk(t)
        if dk not in keep:
            continue
        by_bot_day[(t["bot"], dk)].append(t)
        grp = tag_group(t["bot"])
        for agg, g in AGGREGATE_TAGS.items():
            if grp == g:
                by_bot_day[(agg, dk)].append(t)
    for r in reals:
        dk = _dk(r)
        if dk not in keep or not (rf and dk >= rf):
            continue
        by_real_day[(r["bot"], dk)].append(r)
        grp = tag_group(r["bot"])
        for agg, g in AGGREGATE_TAGS.items():
            if grp == g:
                by_real_day[(agg, dk)].append(r)

    ordered_tags = list(BOT_ORDER)
    seen_tags = {k for k, _ in by_bot_day} | {k for k, _ in by_real_day}
    for extra in sorted(t for t in seen_tags if t not in ordered_tags):
        ordered_tags.append(extra)

    days_out = {}
    for dk in all_days:
        use_reals = bool(rf) and dk >= rf
        rows = []
        for bot in ordered_tags:
            dt_trips = by_bot_day.get((bot, dk), [])
            dt_reals = by_real_day.get((bot, dk), [])
            if not dt_trips and not dt_reals:
                continue
            unit = tag_unit(bot)
            s = summarize(dt_trips, unit, dt_reals if use_reals else None)
            if not s["trades"] and not s["sum_pnl_amt"] and not (s.get("sell_events") or 0):
                continue
            s["key"] = bot
            s["group"] = tag_group(bot)
            s["unit"] = "일" if unit == "day" else ("분" if unit == "min" else "혼합")
            s["day"] = dk
            s["label"] = day_label(dk)
            rows.append(s)
        if rows:
            days_out[dk] = {"label": day_label(dk), "rows": rows}

    return {
        "day_keys": [d for d in all_days if d in days_out],
        "today": date.today().isoformat(),
        "realize_from": rf,
        "days": days_out,
    }


# ───────── 3개월 누적표 ─────────
# 목적: 3개월 단위로 태그별 성적을 줄세워 "뭘 늘리고 뭘 죽일지" 판단.
# 컷오프 리셋 이후 구간만 깨끗하므로, 컷오프가 창 안에 있으면 시작을 컷오프 주로 당긴다.
QUARTER_WEEKS = 13


def build_quarter(by_bot, by_real, all_weeks, ordered_tags, rf):
    """최근 QUARTER_WEEKS 주 구간의 태그별 누적. 주간표와 같은 summarize 를 창 전체에 적용."""
    if not all_weeks:
        return None
    end = all_weeks[-1]
    start = end - timedelta(days=7 * (QUARTER_WEEKS - 1))
    clamped = False
    if rf and start.isoformat() < rf <= end.isoformat():
        start = datetime.strptime(rf, "%Y-%m-%d").date()
        clamped = True
    weeks = [w for w in all_weeks if start <= w <= end]
    if not weeks:
        return None
    wset = {w.isoformat() for w in weeks}
    # 컷오프 이후 주차만 매도시점 실현 — 창이 컷오프를 걸치면 주차별로 갈라 더한다.
    real_weeks = {w for w in wset if rf and w >= rf}
    trip_weeks = wset - real_weeks

    def inwin(x, keys):
        return week_monday(x["exit_dt"].date()).isoformat() in keys

    rows = []
    for tag in ordered_tags:
        bt = [t for t in by_bot.get(tag, []) if inwin(t, wset)]
        amt_src = ([t for t in by_bot.get(tag, []) if inwin(t, trip_weeks)] +
                   [r for r in by_real.get(tag, []) if inwin(r, real_weeks)])
        if not bt and not amt_src:
            continue
        unit = tag_unit(tag)
        s = summarize(bt, unit, amt_src)
        # '매도' 열은 부분청산 회계가 켜진 주차에서만 의미가 있다(그 전 구간은 완결 기준).
        s["sell_events"] = (len([r for r in by_real.get(tag, []) if inwin(r, real_weeks)])
                            if real_weeks else None)
        s["key"] = tag
        s["group"] = tag_group(tag)
        s["unit"] = ("일" if unit == "day" else ("분" if unit == "min" else "혼합"))
        s["label"] = tag
        rows.append(s)

    # 봇(통합 → 태그별 손익순) → 자산배분 → 수동(수동합계 → 태그별) → 레거시
    order = {"bot": 0, "alloc": 1, "manual": 2, "seed": 3, "legacy": 4}
    rows.sort(key=lambda r: (order.get(r["group"], 9),
                             r["key"] not in AGGREGATE_TAGS, -r["sum_pnl_amt"]))
    keys = sorted(wset)
    return {
        "weeks": len(weeks),
        "from": keys[0], "to": keys[-1],
        "label": f"{week_label(weeks[0])[:5]}~{week_label(weeks[-1])[-5:]}",
        "clamped_to_cutoff": clamped,
        "rows": rows,
    }


# ───────── 미국VCP 매수 포착 (buy-only, 실현손익 없음) ─────────
def build_us_vcp(fills, weeks_limit=None):
    """미국 ORDER A(1887US/usvcp) 는 매수만 기록(매도로직 없음)이라 라운드트립이
    안 생긴다 → 표준 성과보드에 안 뜬다. 대신 '매수 포착' 보드로 매수·태그만 보여준다.
    주(월~금)별로 그룹핑, 최근 weeks_limit 주만. 실현손익은 매도데이터 생기기 전까지 없음."""
    buys = [f for f in fills
            if f["acct"] == "1887US" and f["side"] == "buy"
            and (f["strategy"] or "").lower() == "usvcp"]
    if not buys:
        return {"has_data": False, "weeks": [], "total": {"count": 0, "amount": 0}}

    by_week = defaultdict(list)
    for f in buys:
        wk = week_monday(f["dt"].date())
        by_week[wk].append(f)
    all_weeks = sorted(by_week.keys())
    if weeks_limit:
        all_weeks = all_weeks[-weeks_limit:]

    weeks_out = []
    tot_cnt = 0
    tot_amt = 0.0
    for w in sorted(all_weeks, reverse=True):        # 최신 주 먼저
        rows = sorted(by_week[w], key=lambda x: x["dt"])
        wk_rows = []
        wk_amt = 0.0
        for f in rows:
            amt = f["qty"] * f["price"]
            wk_amt += amt
            wk_rows.append({
                "date": f["date"], "code": f["code"], "name": f["name"],
                "qty": f["qty"], "price": round(f["price"], 4),
                "amount": round(amt, 2),
            })
        weeks_out.append({
            "week": w.isoformat(), "label": week_label(w),
            "count": len(wk_rows), "amount": round(wk_amt, 2),
            "buys": wk_rows,
        })
        tot_cnt += len(wk_rows)
        tot_amt += wk_amt
    return {"has_data": True, "weeks": weeks_out,
            "total": {"count": tot_cnt, "amount": round(tot_amt, 2)}}


# ───────── 코인 주간(월~일) 실현손익 ─────────
def _coin_mtf_daily():
    """MTF 본전략 CSV → {(YYYY-MM-DD, 'KRW-BTC'): net_realized}.
    평균단가 replay: 매수는 수수료 포함 원가풀, 매도는 (net_proceeds − 평단×수량).
    오류 manual 행(identifier=manual & uuid 공백, 예: BTC price 1.6억 중복행)은 제외 →
    봇 state json 보유수량과 정확히 일치."""
    daily = defaultdict(float)
    if not os.path.exists(COIN_MTF_CSV):
        return daily
    pos = defaultdict(lambda: {"qty": 0.0, "cost": 0.0})  # 수수료 포함 원가풀
    with open(COIN_MTF_CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            act = row.get("action"); mk = row.get("market")
            if act not in ("BUY", "SELL") or mk not in COIN_MARKETS:
                continue
            # 오류 manual 행 제외(정상 manual 매수는 uuid 존재 → 포함)
            if row.get("identifier") == "manual" and not (row.get("uuid") or "").strip():
                continue
            try:
                vol = float(row.get("volume") or 0); amt = float(row.get("amount_krw") or 0)
            except ValueError:
                continue
            if vol <= 0 or amt <= 0:
                continue
            day = str(row.get("time", ""))[:10]
            p = pos[mk]
            if act == "BUY":
                p["qty"] += vol
                p["cost"] += amt * (1 + COIN_FEE)
            else:  # SELL
                if p["qty"] <= 0:
                    continue  # 대응 매수 없는 매도(CSV 이전 보유분) → 무시
                sq = min(vol, p["qty"])
                avg = p["cost"] / p["qty"]
                net_proceeds = amt * (1 - COIN_FEE) * (sq / vol)
                daily[(day, mk)] += net_proceeds - avg * sq
                p["cost"] -= avg * sq
                p["qty"]  -= sq
    return daily


def _coin_scalp_daily():
    """scalp CSV(v3 우선, 없으면 v2) → {(YYYY-MM-DD, market): realized_pnl 합}.
    scalp 는 realized_pnl 컬럼(봇 자체 집계)을 그대로 매도일에 귀속."""
    daily = defaultdict(float)
    path = next((p for p in COIN_SCALP_CSVS if os.path.exists(p)), None)
    if not path:
        return daily
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            mk = row.get("market")
            if mk not in COIN_MARKETS:
                continue
            try:
                pnl = float(row.get("realized_pnl") or 0)
            except ValueError:
                continue
            if pnl == 0:
                continue
            day = str(row.get("time", ""))[:10]
            daily[(day, mk)] += pnl
    return daily


def build_coin_week(today=None):
    """이번 주(월~일) BTC/ETH 일별 실현손익(net, MTF+scalp) → 웹 table 구조."""
    if today is None:
        today = date.today()
    monday = today - timedelta(days=today.weekday())
    mtf = _coin_mtf_daily()
    scp = _coin_scalp_daily()

    days = []
    total = 0.0
    any_val = False
    for i in range(7):
        d = monday + timedelta(days=i)
        iso = d.isoformat()
        row = {"label": f"{d.month}/{d.day} {COIN_WEEKDAYS[i]}", "btc": None, "eth": None}
        for mk, key in COIN_MARKETS.items():
            has = (iso, mk) in mtf or (iso, mk) in scp
            if has:
                v = int(round(mtf.get((iso, mk), 0.0) + scp.get((iso, mk), 0.0)))
                row[key] = v
                total += v
                any_val = True
        days.append(row)

    fri = monday + timedelta(days=6)
    return {
        "label": f"{monday.month}/{monday.day}~{fri.month}/{fri.day}",
        "total": int(round(total)),
        "has_data": any_val,
        "days": days,
    }


def build_binance_futures_week(today=None):
    """이번 주(월~일) Binance USDT-M 선물 실현손익(net, REALIZED_PNL+COMMISSION+FUNDING_FEE
    합산, USDT) → 웹 table 구조. 원천 = sync_binance_futures_income_v1.py 가 받아온
    income API 기록(0_binance_usdt_futures_income_v1.csv). 심볼이 고정 2개(BTC/ETH)인
    코인과 달리 가변적이라 day row에 등장한 심볼을 동적으로 컬럼화한다."""
    if today is None:
        today = date.today()
    monday = today - timedelta(days=today.weekday())

    daily = defaultdict(lambda: defaultdict(float))  # {iso_date: {symbol: net_usdt}}
    symbols_seen = set()
    if os.path.exists(BINANCE_FUT_INCOME_CSV):
        with open(BINANCE_FUT_INCOME_CSV, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                try:
                    income = float(row.get("income") or 0)
                except ValueError:
                    continue
                if income == 0:
                    continue
                day = str(row.get("time_iso", ""))[:10]
                if not day:
                    continue
                sym = row.get("symbol") or "기타"
                daily[day][sym] += income
                symbols_seen.add(sym)

    symbols_out = sorted(symbols_seen)
    days = []
    total = 0.0
    any_val = False
    for i in range(7):
        d = monday + timedelta(days=i)
        iso = d.isoformat()
        day_data = daily.get(iso, {})
        row = {"label": f"{d.month}/{d.day} {COIN_WEEKDAYS[i]}"}
        day_total = 0.0
        day_has_data = iso in daily
        for sym in symbols_out:
            v = day_data.get(sym)
            row[sym] = round(v, 2) if v else None
            if v:
                day_total += v
        row["total"] = round(day_total, 2) if day_has_data else None
        if day_has_data:
            any_val = True
        total += day_total
        days.append(row)

    sunday = monday + timedelta(days=6)
    return {
        "label": f"{monday.month}/{monday.day}~{sunday.month}/{sunday.day}",
        "unit": "USDT",
        "symbols": symbols_out,
        "total": round(total, 2),
        "has_data": any_val,
        "days": days,
    }


# ───────── 출력 ─────────
def write_json(report, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"  OK {os.path.relpath(path, OUT_DIR)}")


CSV_HEADER = ["봇", "보유단위", "주", "총거래수", "승률(%)", "손익비",
              "평균수익(%)", "평균손실(%)", "최대수익(%)", "최대손실(%)",
              "수익 평균보유", "손실 평균보유",
              "기대값(%/거래)", "기대값(원/거래)", "PF",
              "실현손익합(원,net)", "수수료·세금(원)", "실현손익합(원,gross)",
              "환율(USDKRW)"]


def _week_is_empty(s):
    """이 주차 행을 파일에서 빼도 되는가.

    완결 라운드트립이 0 이어도 부분매도 실현손익은 있을 수 있다(잔량 보유 중).
    trades 만 보고 스킵하면 그 금액이 파일에서 사라지는데 TOTAL 행에는 남아 있어서
    '주차 합 ≠ 전체' 가 된다 — 실제로 2X단타에서 151,458원이 이렇게 증발했다.
    웹 게시판(holdings.html)은 이 행을 그리므로 CSV/xlsx 만 다른 얘기를 하게 된다.
    """
    return not (s.get("trades") or s.get("sum_pnl_amt") or s.get("sum_pnl_gross")
                or s.get("sum_fee_tax") or s.get("sell_events"))


def _row(bot_key, unit, s):
    def g(k):
        v = s.get(k)
        return "" if v is None else v
    return [bot_key, unit, s.get("label", ""), s.get("trades", 0), g("winrate"), g("pl_ratio"),
            g("avg_win"), g("avg_loss"), g("max_win"), g("max_loss"),
            g("hold_win"), g("hold_loss"),
            g("expectancy"), g("expectancy_amt"), g("profit_factor"),
            s.get("sum_pnl_amt", 0), s.get("sum_fee_tax", 0), s.get("sum_pnl_gross", 0),
            g("fx_usdkrw")]   # 전체(TOTAL) 행은 여러 주차가 섞여 공백


def write_csv(report):
    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(CSV_HEADER)
        for b in report["bots"]:
            for wk in b["weeks"]:
                if _week_is_empty(wk):
                    continue
                w.writerow(_row(b["key"], b["unit"], wk))
            w.writerow(_row(b["key"], b["unit"], b["total"]))
            w.writerow([])
    print(f"  OK {os.path.relpath(OUT_CSV, OUT_DIR)}")


def write_xlsx(report):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        print("  (openpyxl 없음 → xlsx 생략, CSV 로 대체)")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "주간성과"
    hdr_fill = PatternFill("solid", fgColor="8B0000")
    hdr_font = Font(color="FFFFFF", bold=True)
    tot_fill = PatternFill("solid", fgColor="F0E0E0")
    ws.append(CSV_HEADER)
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")
    for b in report["bots"]:
        for wk in b["weeks"]:
            if _week_is_empty(wk):
                continue
            ws.append(_row(b["key"], b["unit"], wk))
        ws.append(_row(b["key"], b["unit"], b["total"]))
        for c in ws[ws.max_row]:
            c.fill = tot_fill; c.font = Font(bold=True)
        ws.append([])
    widths = [14, 9, 14, 9, 8, 8, 11, 11, 11, 11, 12, 12, 13, 14, 7, 18, 15, 18, 13]
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = wd
    wb.save(OUT_XLSX)
    print(f"  OK {os.path.relpath(OUT_XLSX, OUT_DIR)}")


def _vw(s):
    """콘솔 표시 폭 (CJK/이모지 = 2칸)."""
    import unicodedata
    return sum(2 if unicodedata.east_asian_width(c) in ("W", "F") else 1 for c in str(s))


def _pad(s, width, right=False):
    s = str(s)
    gap = " " * max(width - _vw(s), 0)
    return (gap + s) if right else (s + gap)


def print_bot_summary(report, indent="  ", show_zero=True):
    """봇별 전체기간 요약표. 한글 봇이름이 섞여도 열이 맞도록 표시폭 기준 정렬."""
    rows = []
    for b in report["bots"]:
        t = b["total"]
        rows.append({
            "group":  b.get("group", "bot"),
            "key":    b["key"],
            "trades": t["trades"],
            "wr":     "-" if t["winrate"] is None else f"{t['winrate']}%",
            "ex":     "-" if t["expectancy"] is None else f"{t['expectancy']:+.2f}%",
            "pf":     "-" if t["profit_factor"] is None else f"{t['profit_factor']}",
            "net":    f"{t['sum_pnl_amt']:,}원",
            "fee":    f"{t['sum_fee_tax']:,}원",
        })
    live = [r for r in rows if r["trades"] > 0]
    idle = [r for r in rows if r["trades"] == 0]

    w_key = max([_vw(r["key"]) for r in rows] + [_vw("봇")])
    w_net = max([_vw(r["net"]) for r in live] + [_vw("순손익")])
    w_fee = max([_vw(r["fee"]) for r in live] + [_vw("비용")])
    head = (indent + _pad("봇", w_key) + "  " + _pad("거래", 5, True) + "  " + _pad("승률", 7, True)
            + "  " + _pad("기대값", 8, True) + "  " + _pad("PF", 6, True)
            + "  " + _pad("순손익", w_net, True) + "  " + _pad("비용", w_fee, True))
    line = indent + "-" * (_vw(head) - _vw(indent))

    print(f"\n{indent}[태그별 전체기간 요약]  (웹 표시=최근 {RECENT_WEEKS}주)")
    print(f"{indent}'통합'=봇 태그만 합산. 자산배분(통합ETF)·수동은 별도 그룹.")
    print(line)
    print(head)
    print(line)
    grp_label = {"bot": "[봇]", "alloc": "[자산배분]", "manual": "[수동 — 통합 미포함]",
                 "legacy": "[레거시 편입분 — 통합 미포함]"}
    prev_grp = None
    for r in live:
        if r["group"] != prev_grp:
            if prev_grp is not None:
                print(line)
            print(indent + grp_label.get(r["group"], r["group"]))
            prev_grp = r["group"]
        print(indent + _pad(r["key"], w_key) + "  " + _pad(r["trades"], 5, True)
              + "  " + _pad(r["wr"], 7, True) + "  " + _pad(r["ex"], 8, True)
              + "  " + _pad(r["pf"], 6, True)
              + "  " + _pad(r["net"], w_net, True)
              + "  " + _pad(r["fee"], w_fee, True))
    print(line)
    if idle and show_zero:
        print(f"{indent}거래 0건: " + ", ".join(r["key"] for r in idle))


def print_day_summary(report, day=None, indent="  "):
    """하루치 태그별 성과. day 미지정이면 오늘. daily 블록이 없는 예전 json 이면 조용히 통과.

    '거래'는 완결 라운드트립, '매도'는 부분청산 포함 매도 건수다. 두 값이 다른 게 정상이고
    (예: 거래 0 · 매도 3 = 아직 완결 안 된 포지션의 분할매도), 금액은 매도 기준이라
    거래 0 인 줄에도 손익이 찍힌다.
    """
    daily = (report or {}).get("daily")
    if not daily:
        return
    dk = day or daily.get("today") or date.today().isoformat()
    entry = (daily.get("days") or {}).get(dk)

    print(f"\n{indent}[{day_label(dk)} 당일]  거래=완결 · 매도=부분청산 포함")
    if not entry:
        print(f"{indent}(완결·실현 없음)")
        return

    rows = []
    for r in entry["rows"]:
        rows.append({
            "group":  r.get("group", "bot"),
            "key":    r["key"],
            "trades": r["trades"],
            "sells":  r.get("sell_events") or 0,
            "wr":     "-" if r["winrate"] is None else f"{r['winrate']}%",
            "net":    f"{r['sum_pnl_amt']:,}원",
            "fee":    f"{r['sum_fee_tax']:,}원",
        })
    w_key = max([_vw(r["key"]) for r in rows] + [_vw("봇")])
    w_net = max([_vw(r["net"]) for r in rows] + [_vw("순손익")])
    w_fee = max([_vw(r["fee"]) for r in rows] + [_vw("비용")])
    head = (indent + _pad("봇", w_key) + "  " + _pad("거래", 5, True) + "  " + _pad("매도", 5, True)
            + "  " + _pad("승률", 7, True)
            + "  " + _pad("순손익", w_net, True) + "  " + _pad("비용", w_fee, True))
    line = indent + "-" * (_vw(head) - _vw(indent))
    grp_label = {"bot": "[봇]", "alloc": "[자산배분]", "manual": "[수동 — 통합 미포함]",
                 "legacy": "[레거시 편입분 — 통합 미포함]"}
    print(line)
    print(head)
    print(line)
    prev_grp = None
    for r in rows:
        if r["group"] != prev_grp:
            if prev_grp is not None:
                print(line)
            print(indent + grp_label.get(r["group"], r["group"]))
            prev_grp = r["group"]
        print(indent + _pad(r["key"], w_key) + "  " + _pad(r["trades"], 5, True)
              + "  " + _pad(r["sells"], 5, True) + "  " + _pad(r["wr"], 7, True)
              + "  " + _pad(r["net"], w_net, True) + "  " + _pad(r["fee"], w_fee, True))
    print(line)


def main():
    print("── 주간성과 집계 (intraday + Pine TR tag 기반) ──")
    load_fx_snapshot()   # 과거 주차 환율 동결값 로드 (원화 금액 재현성)
    fills = load_fills()
    print(f"  fill 이벤트: {len(fills)}건")
    intraday_trips = build_round_trips(fills)
    tr_trips = build_tr_ledger_round_trips()
    trips = intraday_trips + tr_trips
    print(f"  intraday 완결 라운드트립: {len(intraday_trips)}건")
    print(f"  TR ledger 완결 라운드트립: {len(tr_trips)}건")
    print(f"  전체 완결 라운드트립: {len(trips)}건")
    drift = build_drift_payload(check_position_drift())

    # 전체 누적(분석용): full json + csv + xlsx
    full_report = build_report(trips)
    full_report["drift"] = drift
    # 일별 버킷은 full json 에만 붙인다(웹 피드 weekly_performance.json 은 그대로).
    # csv/xlsx 는 report["bots"] 만 훑으므로 최상위 키가 늘어도 영향 없다.
    full_report["daily"] = build_daily_report(trips)
    d = full_report["daily"]
    print(f"  일별 버킷: 활동일 {len(d['day_keys'])}일"
          + (f" (오늘 {d['today']} 포함)" if d["today"] in d["days"] else f" (오늘 {d['today']} 활동없음)"))
    write_json(full_report, OUT_JSON_FULL)
    write_csv(full_report)
    write_xlsx(full_report)

    # 웹페이지 피드: 최근 N주만 (총계도 그 구간 기준)
    web_report = build_report(trips, weeks_limit=RECENT_WEEKS)
    # 잔고 드리프트 경고: 지금까지 콘솔에만 찍혀서 아무도 못 봤다 → 게시판 상단 배너로.
    web_report["drift"] = drift
    # 3개월 누적표는 웹 표시구간(8주)이 아니라 전체 데이터에서 뽑은 걸 쓴다.
    web_report["quarter"] = full_report["quarter"]
    q = web_report["quarter"]
    if q:
        print(f"  3개월 누적표: {q['label']} ({q['weeks']}주, 태그 {len(q['rows'])}개"
              + (", 컷오프 이후로 제한" if q["clamped_to_cutoff"] else "") + ")")
    # 코인(업비트) 이번 주(월~일) 실현손익 미니표 데이터 부착
    web_report["coin_week"] = build_coin_week()
    cw = web_report["coin_week"]
    print(f"  코인 이번주({cw['label']}) 실현손익 합: {cw['total']:,}원")
    # 미국VCP 매수 포착(buy-only) 보드 데이터 부착
    web_report["us_vcp"] = build_us_vcp(fills, weeks_limit=RECENT_WEEKS)
    uv = web_report["us_vcp"]
    print(f"  미국VCP 매수 포착: {uv['total']['count']}건 (${uv['total']['amount']:,.0f})")
    # Binance USDT-M 선물 이번 주(월~일) 실현손익 미니표 데이터 부착
    web_report["binance_futures_week"] = build_binance_futures_week()
    bf = web_report["binance_futures_week"]
    print(f"  Binance선물 이번주({bf['label']}) 실현손익 합: {bf['total']:,.2f} USDT")
    write_json(web_report, OUT_JSON)
    save_fx_snapshot()   # 이번 실행에서 쓴 주차별 환율 기록 → 다음 실행부터 고정

    # 콘솔 요약(전체기간)
    print_bot_summary(full_report)


if __name__ == "__main__":
    main()

